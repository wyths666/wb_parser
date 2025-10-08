import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink


def redact(df, name = 'wb_products_formatted.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = "WB Products"

# Заголовки
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

# Добавляем заголовки
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Цвета для чередования строк
    fill_odd = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # Светло-серый
    fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Белый

# Добавляем данные
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):  # Начинаем с 2-й строки
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)

        # Применяем чередование фона
            if row_num % 2 == 0:
                cell.fill = fill_even
            else:
                cell.fill = fill_odd

        # Форматирование цены (только для столбца 'price')
            if df.columns[col_num - 1] == 'price':
                cell.number_format = '#,##0 ₽'

        # Гиперссылки для 'product_url' и 'photo_url'
            col_name = df.columns[col_num - 1]
            if col_name in ['product_url', 'photo_url'] and isinstance(value, str) and value.startswith('http'):
                cell.hyperlink = value
                cell.style = "Hyperlink"

# Автоподбор ширины столбцов

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


    column_names = df.columns.tolist()
    for col_idx, col_name in enumerate(column_names, start=1):
        if col_name in ['product_url', 'photo_url']:
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15
        if col_name == 'price':
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 11
        if col_name == 'full_name':
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 70
        #full_name

# Сохраняем файл

    output_file = name
    wb.save(output_file)
    print(f"Файл успешно сохранён: {output_file}")